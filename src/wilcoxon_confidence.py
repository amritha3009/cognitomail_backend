import os
import sys
import numpy as np

print("Script started...")
print("Current folder:", os.getcwd())

# --- try to import required packages ---
try:
    from scipy.stats import wilcoxon
    import openpyxl
    print("scipy and openpyxl loaded OK")
except ImportError as e:
    print("Missing package:", e)
    print("Run:  py -m pip install scipy numpy openpyxl")
    sys.exit(1)

# --- path to your Excel file ---
# Option A: file is in the same folder as the script
excel_path = "Survey output.xlsx"

# Option B: if the file is somewhere else, use the full path, e.g.:
# excel_path = r"C:\Users\amrit\Downloads\Survey output.xlsx"

if not os.path.exists(excel_path):
    print("ERROR: Excel file not found at:", excel_path)
    print("Files in this folder:")
    for f in os.listdir("."):
        print("  ", f)
    sys.exit(1)

print("Opening:", excel_path)
wb = openpyxl.load_workbook(excel_path, data_only=True)
print("Sheet names:", wb.sheetnames)

# Use the exact sheet name from your file
sheet_name = "Post-Study Survey - CognitoMail"
if sheet_name not in wb.sheetnames:
    print("ERROR: Sheet not found. Available sheets:", wb.sheetnames)
    sys.exit(1)

ws = wb[sheet_name]
print("Max row:", ws.max_row)

pairs = []

# Row 2 – columns are aligned
try:
    r1 = int(float(ws.cell(2, 5).value))
    r2 = int(float(ws.cell(2, 8).value))
    pairs.append((r1, r2))
    print("Row 2 OK:", r1, r2)
except Exception as e:
    print("Row 2 problem:", e)

# Rows 3 onwards – SurveyMonkey shifted columns
for r in range(3, ws.max_row + 1):
    v1 = ws.cell(r, 14).value
    v2 = ws.cell(r, 17).value
    if v1 is None or v2 is None:
        continue
    try:
        r1 = int(float(v1))
        r2 = int(float(v2))
        if 1 <= r1 <= 5 and 1 <= r2 <= 5:
            pairs.append((r1, r2))
    except:
        continue

print("Number of pairs found:", len(pairs))

if len(pairs) < 5:
    print("Too few pairs – check column numbers in the Excel file.")
    sys.exit(1)

round1 = np.array([p[0] for p in pairs], dtype=float)
round2 = np.array([p[1] for p in pairs], dtype=float)

print(f"Round1 mean = {round1.mean():.4f}")
print(f"Round2 mean = {round2.mean():.4f}")

stat, p_one = wilcoxon(round2, round1, alternative="greater")
stat2, p_two = wilcoxon(round2, round1, alternative="two-sided")

diffs = round2 - round1
n_eff = int(np.sum(diffs != 0))
total = n_eff * (n_eff + 1) / 2
pos = float(stat)
neg = total - pos
r_rb = (pos - neg) / total if total > 0 else 0

print("--- Wilcoxon results ---")
print(f"n = {len(pairs)}")
print(f"Positive rank sum = {stat}")
print(f"p (one-sided) = {p_one:.6f}")
print(f"p (two-sided) = {p_two:.6f}")
print(f"Rank-biserial r = {r_rb:.4f}")
print("Script finished successfully.")